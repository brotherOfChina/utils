import qrcode
from PIL import Image, ImageDraw, ImageFont
import os
import numpy
import cv2
import openpyxl
from docx import Document
from docx.shared import Inches

doc = Document()


def save_img_to_word(img_path):  # 保存图片到word
    print("save_img_to_word")
    doc.add_picture(img_path, width=Inches(6))  # 添加图, 设置宽度
    doc.save('egg.docx')


# 添加图片到原有图片
def add_qr_code_to_img(icon, img_path):
    print("add_qr_code_to_img")
    # icon = Image.open(icon_path)
    icon.resize((21, 21))
    img = Image.open(img_path)
    img.paste(icon, (10, 10), mask=None)
    img.save("I:/code/code.jpg")
    save_img_to_word("I:/code/code.jpg")


def create_qr_code(content, path, num):  # 生成带有num的二维码
    print("create_qr_code{}:{}".format(content, num))
    qr = qrcode.QRCode(
        version=2,  # 生成二维码尺寸的大小 1-40  1:21*21（21+(n-1)*4）  4*12*10
        error_correction=qrcode.constants.ERROR_CORRECT_M,  # L:7% M:15% Q:25% H:30%
        box_size=10,  # 每个格子的像素大小
        border=2,  # 边框的格子宽度大小
    )
    padding = 5
    qr.add_data(data=content)
    qr.make(fit=True)
    img = qr.make_image()
    font = ImageFont.truetype('3.otf', 35)
    w, h = font.getsize("{}".format(num))
    icon = Image.new("RGB", (w + padding * 2, h + padding * 2))
    draw = ImageDraw.Draw(icon)
    draw.text(text="{}".format(num), xy=(padding, 0), font=font)
    icon = icon.resize((w + padding * 2, h + padding * 2), Image.ANTIALIAS)
    img.paste(icon, (120, 140), mask=None)
    img.save("I:/code/{}.png".format(num))
    add_qr_code_to_img(img, "I:/code/egg.jpg")


def read_excel(path):  # 读取指定路径下的xlsx
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    columA = ws["A"]
    maxrow = ws.max_row
    for index in range(maxrow):
        print(columA[index].value)
        create_qr_code(columA[index].value, "", index)


read_excel("I:/code/codes.xlsx")

# save_img_to_word("I:/code/egg.jpg")
